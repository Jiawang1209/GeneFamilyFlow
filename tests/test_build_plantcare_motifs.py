"""Tests for scripts/build_plantcare_motifs.py.

Covers:
- PlantCARE .tab parsing with element-whitelist filtering.
- Rejection of non-DNA sequences, ``Unname`` elements, and
  ``short_function`` rows.
- Merge semantics: existing rows preserved, new pairs appended, entries
  sorted by (description, element, sequence).
- CLI round-trip (``--from-plantcare-tab`` + ``--elements-xlsx``).
- Shipped fixture integrity: example/10.promoter/plantcare_motifs.tsv is
  non-empty, every row matches a xlsx element, and every sequence is DNA.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from scripts.build_plantcare_motifs import (
    Motif,
    _read_element_descriptions,
    main,
    merge_motifs,
    parse_plantcare_tab,
    read_existing,
    write_tsv,
)

pytestmark = pytest.mark.unit

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURE_TSV = REPO_ROOT / "example" / "10.promoter" / "plantcare_motifs.tsv"
FIXTURE_XLSX = (
    REPO_ROOT / "example" / "10.promoter" / "cir_element.desc.20240509.xlsx"
)


def _write_xlsx(tmp_path: Path, rows: list[tuple[str, str]]) -> Path:
    path = tmp_path / "elements.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["element", "description"])
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    return path


def _write_tab(tmp_path: Path, name: str, rows: list[tuple[str, ...]]) -> Path:
    path = tmp_path / name
    with path.open("w") as fh:
        for r in rows:
            fh.write("\t".join(r) + "\n")
    return path


class TestReadElementDescriptions:
    def test_basic(self, tmp_path: Path) -> None:
        xlsx = _write_xlsx(
            tmp_path,
            [("ABRE", "Plant hormone related"), ("CAAT-box", "Promoter")],
        )
        assert _read_element_descriptions(xlsx) == {
            "ABRE": "Plant hormone related",
            "CAAT-box": "Promoter",
        }

    def test_missing_columns_raises(self, tmp_path: Path) -> None:
        path = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        wb.active.append(["foo", "bar"])
        wb.save(path)
        with pytest.raises(ValueError, match="element"):
            _read_element_descriptions(path)

    def test_skips_blank_rows(self, tmp_path: Path) -> None:
        xlsx = _write_xlsx(
            tmp_path,
            [("ABRE", "Hormone"), ("", "Light"), ("Empty", "")],
        )
        assert _read_element_descriptions(xlsx) == {"ABRE": "Hormone"}


class TestParsePlantcareTab:
    def test_filters_to_whitelist(self, tmp_path: Path) -> None:
        tab = _write_tab(
            tmp_path,
            "a.tab",
            [
                ("gene1", "ABRE", "ACGTG", "54", "5", "-", "Ath", "abscisic acid"),
                ("gene1", "CAAT-box", "CAAAT", "88", "5", "+", "Psat", "common"),
                ("gene1", "NotInXlsx", "AAAA", "1", "4", "+", "Ath", "nope"),
            ],
        )
        allowed = {"ABRE": "Hormone", "CAAT-box": "Promoter"}
        motifs = parse_plantcare_tab([tab], allowed)
        assert sorted((m.element, m.sequence) for m in motifs) == [
            ("ABRE", "ACGTG"),
            ("CAAT-box", "CAAAT"),
        ]
        for m in motifs:
            assert m.description == allowed[m.element]

    def test_rejects_non_dna_and_short_function(self, tmp_path: Path) -> None:
        tab = _write_tab(
            tmp_path,
            "a.tab",
            [
                ("g1", "ABRE", "ACXGT", "1", "5", "+", "Ath", "ok"),          # non-DNA
                ("g1", "ABRE", "", "1", "5", "+", "Ath", "ok"),                # empty seq
                ("g1", "ABRE", "ACGTG", "1", "5", "+", "Ath", "short_function"),
                ("g1", "ABRE", "TTTTT", "1", "5", "+", "Ath", ""),             # empty ok
                ("g1", "ABRE", "ACGTG", "1", "5", "+", "Ath", "abscisic"),
            ],
        )
        motifs = parse_plantcare_tab([tab], {"ABRE": "Hormone"})
        assert sorted(m.sequence for m in motifs) == ["ACGTG", "TTTTT"]

    def test_dedupe_across_files(self, tmp_path: Path) -> None:
        tab_a = _write_tab(tmp_path, "a.tab", [
            ("g1", "ABRE", "ACGTG", "1", "5", "+", "Ath", "one"),
        ])
        tab_b = _write_tab(tmp_path, "b.tab", [
            ("g2", "ABRE", "ACGTG", "2", "5", "+", "Osat", "two"),  # duplicate pair
            ("g2", "ABRE", "CACGTG", "2", "6", "+", "Ath", "three"),
        ])
        motifs = parse_plantcare_tab([tab_a, tab_b], {"ABRE": "Hormone"})
        assert len(motifs) == 2
        seqs = sorted(m.sequence for m in motifs)
        assert seqs == ["ACGTG", "CACGTG"]

    def test_lowercases_to_uppercase(self, tmp_path: Path) -> None:
        tab = _write_tab(tmp_path, "a.tab", [
            ("g1", "ABRE", "acgtg", "1", "5", "+", "Ath", "ok"),
        ])
        motifs = parse_plantcare_tab([tab], {"ABRE": "Hormone"})
        assert motifs[0].sequence == "ACGTG"


class TestMergeMotifs:
    def test_preserves_existing(self) -> None:
        existing = {
            ("ABRE", "ACGTG"): Motif("ABRE", "ACGTG", "H", "user", "keep me"),
        }
        additions = [Motif("ABRE", "ACGTG", "H", "override", "replace me")]
        merged = merge_motifs(existing, additions)
        assert len(merged) == 1
        assert merged[0].function_desc == "keep me"

    def test_appends_new_pairs_sorted(self) -> None:
        existing = {
            ("ABRE", "ACGTG"): Motif("ABRE", "ACGTG", "Hormone", "a", "a"),
        }
        additions = [
            Motif("Box4", "ATTAAT", "Light", "b", "b"),
            Motif("CAAT-box", "CAAAT", "Promoter", "c", "c"),
        ]
        merged = merge_motifs(existing, additions)
        # Sorted by (description, element, sequence)
        assert [m.element for m in merged] == ["ABRE", "Box4", "CAAT-box"]


class TestReadExistingRoundtrip:
    def test_roundtrip(self, tmp_path: Path) -> None:
        p = tmp_path / "motifs.tsv"
        motifs = [
            Motif("ABRE", "ACGTG", "Hormone", "Ath", "abscisic"),
            Motif("CAAT-box", "CAAAT", "Promoter", "Psat", "common"),
        ]
        write_tsv(motifs, p)
        back = read_existing(p)
        assert len(back) == 2
        assert back[("ABRE", "ACGTG")].function_desc == "abscisic"

    def test_missing_file_returns_empty(self, tmp_path: Path) -> None:
        assert read_existing(tmp_path / "nope.tsv") == {}

    def test_bad_header_raises(self, tmp_path: Path) -> None:
        p = tmp_path / "bad.tsv"
        p.write_text("a\tb\tc\td\te\nfoo\tbar\tbaz\tqux\tquux\n")
        with pytest.raises(ValueError, match="header"):
            read_existing(p)


class TestCLI:
    def test_roundtrip(self, tmp_path: Path) -> None:
        xlsx = _write_xlsx(tmp_path, [("ABRE", "Hormone"), ("CAAT-box", "Promoter")])
        tab = _write_tab(
            tmp_path,
            "a.tab",
            [
                ("g1", "ABRE", "ACGTG", "1", "5", "+", "Ath", "abscisic"),
                ("g1", "CAAT-box", "CAAAT", "1", "5", "+", "Psat", "common"),
                ("g1", "NotListed", "TTTT", "1", "4", "+", "Ath", "nope"),
            ],
        )
        out = tmp_path / "out.tsv"
        rc = main([
            "--from-plantcare-tab", str(tab),
            "--elements-xlsx", str(xlsx),
            "-o", str(out),
        ])
        assert rc == 0
        body = out.read_text().splitlines()
        assert body[0].split("\t") == [
            "element", "sequence", "description", "species", "function_desc"
        ]
        elements = [line.split("\t")[0] for line in body[1:]]
        assert sorted(elements) == ["ABRE", "CAAT-box"]

    def test_merge_preserves_prior(self, tmp_path: Path) -> None:
        xlsx = _write_xlsx(tmp_path, [("ABRE", "Hormone"), ("CAAT-box", "Promoter")])

        seed_tab = _write_tab(
            tmp_path, "seed.tab",
            [("g1", "ABRE", "ACGTG", "1", "5", "+", "user", "seeded row")],
        )
        base = tmp_path / "base.tsv"
        main([
            "--from-plantcare-tab", str(seed_tab),
            "--elements-xlsx", str(xlsx),
            "-o", str(base),
        ])

        new_tab = _write_tab(
            tmp_path, "new.tab",
            [
                ("g2", "ABRE", "ACGTG", "1", "5", "+", "override", "should not win"),
                ("g2", "CAAT-box", "CAAAT", "1", "5", "+", "Psat", "common"),
            ],
        )
        out = tmp_path / "out.tsv"
        rc = main([
            "--from-plantcare-tab", str(new_tab),
            "--elements-xlsx", str(xlsx),
            "--merge", str(base),
            "-o", str(out),
        ])
        assert rc == 0
        back = read_existing(out)
        assert back[("ABRE", "ACGTG")].function_desc == "seeded row"
        assert ("CAAT-box", "CAAAT") in back


class TestShippedFixture:
    def test_fixture_exists_and_parses(self) -> None:
        assert FIXTURE_TSV.exists(), f"missing fixture {FIXTURE_TSV}"
        entries = read_existing(FIXTURE_TSV)
        assert len(entries) > 50, "fixture should ship at least 50 motifs"

    def test_all_elements_in_xlsx(self) -> None:
        allowed = set(_read_element_descriptions(FIXTURE_XLSX))
        entries = read_existing(FIXTURE_TSV)
        for (element, _seq) in entries:
            assert element in allowed, f"fixture element not in xlsx: {element}"

    def test_sequences_are_dna(self) -> None:
        entries = read_existing(FIXTURE_TSV)
        for (_elem, seq) in entries:
            assert seq, "empty sequence"
            assert all(b in "ACGTN" for b in seq), f"non-DNA base in {seq}"

    def test_covers_all_four_categories(self) -> None:
        cats = {m.description for m in read_existing(FIXTURE_TSV).values()}
        assert cats == {
            "Light responsiveness",
            "Plant growth and development",
            "Plant hormone related",
            "Stress related",
        }


class TestDefensiveBranches:
    """Exercise the remaining defensive skips so coverage reflects intent."""

    def test_read_element_descriptions_skips_empty_row(self, tmp_path: Path) -> None:
        """openpyxl may yield an empty tuple for fully-blank rows (line 73)."""
        path = tmp_path / "blank.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["element", "description"])
        ws.append(["ABRE", "Hormone"])
        ws.append([])  # fully blank row
        ws.append(["CAAT-box", "Promoter"])
        wb.save(path)
        assert _read_element_descriptions(path) == {
            "ABRE": "Hormone",
            "CAAT-box": "Promoter",
        }

    def test_parse_plantcare_tab_skips_blank_and_truncated(
        self, tmp_path: Path
    ) -> None:
        """Blank lines (line 99) and <8-column rows (line 102) are dropped."""
        path = tmp_path / "mixed.tab"
        path.write_text(
            "\n"  # blank line
            "g1\tABRE\tonly_three_cols\n"  # truncated row
            "g1\tABRE\tACGTG\t1\t5\t+\tAth\tok\n"
        )
        motifs = parse_plantcare_tab([path], {"ABRE": "Hormone"})
        assert len(motifs) == 1
        assert motifs[0].sequence == "ACGTG"

    def test_read_existing_skips_blank_and_malformed_rows(
        self, tmp_path: Path
    ) -> None:
        """Blank lines (143), wrong column count (146), and empty element/seq
        rows (149) are silently dropped when reading a prior motif library."""
        path = tmp_path / "library.tsv"
        path.write_text(
            "element\tsequence\tdescription\tspecies\tfunction_desc\n"
            "ABRE\tACGTG\tHormone\tAth\tabscisic\n"
            "\n"  # blank line → 143
            "only\tfour\tcolumns\there\n"  # 4 cols → 146
            "\t\tHormone\tAth\tabscisic\n"  # empty element & sequence → 149
            "CAAT-box\tCAAAT\tPromoter\tPsat\tcommon\n"
        )
        out = read_existing(path)
        assert set(out) == {("ABRE", "ACGTG"), ("CAAT-box", "CAAAT")}

    def test_module_entrypoint(self, tmp_path: Path) -> None:
        """``python -m scripts.build_plantcare_motifs`` exercises the
        ``raise SystemExit(main())`` tail (line 226)."""
        import runpy

        xlsx = _write_xlsx(tmp_path, [("ABRE", "Hormone")])
        tab = _write_tab(
            tmp_path, "a.tab",
            [("g1", "ABRE", "ACGTG", "1", "5", "+", "Ath", "abscisic")],
        )
        out = tmp_path / "out.tsv"

        import sys
        original = sys.argv
        sys.argv = [
            "build_plantcare_motifs.py",
            "--from-plantcare-tab", str(tab),
            "--elements-xlsx", str(xlsx),
            "-o", str(out),
        ]
        try:
            with pytest.raises(SystemExit) as exc:
                runpy.run_module(
                    "scripts.build_plantcare_motifs", run_name="__main__"
                )
            assert exc.value.code == 0
        finally:
            sys.argv = original
        assert out.exists()
